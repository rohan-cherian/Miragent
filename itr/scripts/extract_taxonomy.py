"""
Task 0 — extract the real taxonomy from the workbook into data/taxonomy.json.

One-time (well, re-runnable), deterministic. Reads
data/ITR_SyntheticData_GenerationClasses_v1_30July2026.xlsx with openpyxl and
writes data/taxonomy.json — the single source of truth scripts/seed_taxonomy.py
loads from. Never talks to Postgres; pure file-in, file-out.

Sheets, exactly (verified by direct inspection of the workbook):
  04_L1_Categories  — header row 3, data rows 4-13, 10 categories.
                      Columns: Code, Category, Definition, Corpus share.
  05_Problem_Classes — header row 3, data rows 4-103, 100 classes, 10 per
                      category. 15 columns: Class ID, Problem class,
                      Description, Entry tier, Complexity (1-5), Avalara fit,
                      Stripe fit, Generic SaaS fit, Primary skill,
                      Default priority, P50 time-to-resolve (h),
                      Escalation rate, ITR disposition, KB article exists,
                      CSAT risk.

Example-phrases overlay
------------------------
The workbook has no phrases column. Hand-written phrases live in
data/example_phrases.json (`{"CFG-09": ["...", ...]}`) and are MERGED in
here — re-running extraction never destroys hand-written work, because the
overlay file is read fresh every time and is the only source of phrases.
A class absent from the overlay gets `example_phrases: []`.

default_priority is kept RAW here (low | normal | high | urgent, whatever the
workbook says) — the low/medium/high/critical mapping onto triage.py's ladder
happens at seed time in scripts/seed_taxonomy.py, not here, so this file
stays a faithful transcript of the workbook.

Usage:
  poetry run python scripts/extract_taxonomy.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl

WORKBOOK_PATH = ROOT / "data" / "ITR_SyntheticData_GenerationClasses_v1_30July2026.xlsx"
PHRASES_PATH = ROOT / "data" / "example_phrases.json"
OUTPUT_PATH = ROOT / "data" / "taxonomy.json"

CATEGORIES_SHEET = "04_L1_Categories"
CATEGORIES_FIRST_ROW = 4
CATEGORIES_LAST_ROW = 13

CLASSES_SHEET = "05_Problem_Classes"
CLASSES_FIRST_ROW = 4
CLASSES_LAST_ROW = 103

EXPECTED_CATEGORY_COUNT = 10
EXPECTED_CLASS_COUNT = 100


def _load_phrases() -> dict[str, list[str]]:
    if not PHRASES_PATH.exists():
        return {}
    return json.loads(PHRASES_PATH.read_text(encoding="utf-8"))


def _extract_categories(workbook) -> list[dict]:
    sheet = workbook[CATEGORIES_SHEET]
    categories = []
    for row in range(CATEGORIES_FIRST_ROW, CATEGORIES_LAST_ROW + 1):
        code, name, definition, corpus_share = (sheet.cell(row=row, column=c).value for c in range(1, 5))
        if not code:
            continue
        categories.append(
            {
                "code": str(code).strip(),
                "name": str(name).strip() if name else "",
                "definition": str(definition).strip() if definition else "",
                "corpus_share": corpus_share,
            }
        )
    return categories


def _extract_classes(workbook, phrases: dict[str, list[str]]) -> list[dict]:
    sheet = workbook[CLASSES_SHEET]
    classes = []
    for row in range(CLASSES_FIRST_ROW, CLASSES_LAST_ROW + 1):
        values = [sheet.cell(row=row, column=c).value for c in range(1, 16)]
        (
            class_id,
            name,
            description,
            entry_tier,
            complexity,
            _avalara_fit,
            _stripe_fit,
            _generic_saas_fit,
            _primary_skill,
            default_priority,
            _p50_hours,
            _escalation_rate,
            disposition,
            kb_article_exists,
            csat_risk,
        ) = values

        if not class_id:
            continue

        class_id = str(class_id).strip()
        classes.append(
            {
                "class_id": class_id,
                "category": class_id[:3],
                "name": str(name).strip() if name else "",
                "description": str(description).strip() if description else "",
                "entry_tier": str(entry_tier).strip() if entry_tier else None,
                "complexity": complexity,
                "default_priority": str(default_priority).strip().lower() if default_priority else None,
                "disposition": str(disposition).strip() if disposition else None,
                "kb_article_exists": str(kb_article_exists).strip().upper() == "Y",
                "csat_risk": str(csat_risk).strip() if csat_risk else None,
                "example_phrases": list(phrases.get(class_id, [])),
            }
        )
    return classes


def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"FATAL: workbook not found at {WORKBOOK_PATH}")
        return 1

    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    phrases = _load_phrases()

    categories = _extract_categories(workbook)
    classes = _extract_classes(workbook, phrases)

    known_codes = {c["code"] for c in categories}
    per_category: dict[str, int] = {}
    unknown_prefix_classes = []
    with_phrases = 0

    for entry in classes:
        per_category[entry["category"]] = per_category.get(entry["category"], 0) + 1
        if entry["category"] not in known_codes:
            unknown_prefix_classes.append(entry["class_id"])
        if entry["example_phrases"]:
            with_phrases += 1

    print(f"categories: {len(categories)}")
    print(f"classes:    {len(classes)}")
    print("per-category class counts:")
    for code in sorted(per_category):
        print(f"  {code}: {per_category[code]}")
    if unknown_prefix_classes:
        print(f"classes with an unrecognised category prefix: {unknown_prefix_classes}")
    print(f"classes with example_phrases: {with_phrases}/{len(classes)}")

    if len(categories) != EXPECTED_CATEGORY_COUNT or len(classes) != EXPECTED_CLASS_COUNT:
        print(
            f"\nFATAL: expected exactly {EXPECTED_CATEGORY_COUNT} categories and "
            f"{EXPECTED_CLASS_COUNT} classes, got {len(categories)} and {len(classes)} — "
            "a silently short extraction is how bad label spaces ship."
        )
        return 1

    output = {"categories": categories, "classes": classes}
    OUTPUT_PATH.write_text(json.dumps(output, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"\nwrote {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
